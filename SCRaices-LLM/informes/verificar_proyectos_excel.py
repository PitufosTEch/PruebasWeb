"""
verificar_proyectos_excel.py
============================
Compara los proyectos activos en el sistema de curvas S con las pestañas
existentes en Proyeccion_Despachos_2026.xlsx. Para cada proyecto activo
sin pestaña, crea una hoja en blanco con el formato estándar.

Se ejecuta diariamente (GitHub Actions) para detectar automáticamente
proyectos nuevos agregados al dashboard, sin necesidad de intervención manual.

Fuente de verdad: PROYECTOS_ACTIVOS (debe espejarse con curvas_semanal.yml).
"""

import io
import sys
from pathlib import Path

# Lista maestra de proyectos activos — mantener sincronizada con curvas_semanal.yml
# Al agregar un proyecto nuevo al workflow de curvas S, agregar aquí también.
PROYECTOS_ACTIVOS = [
    ("P119", "Nuke Mapu"),
    ("P38",  "Aliwen"),
    ("P39",  "El Coihue"),
    ("P127", "Nuevo Cunco"),
    ("P12",  "Juan Huilcan Tolten"),
    ("P14",  "Com. Madihue"),
    ("P126", "El Maiten"),
    ("P131", "Raices Melipeuco"),
    ("P116", "Sonia Quilaleo"),
    ("P31",  "Trovolhue"),
    ("P28",  "Elsa Pinchulaf"),
]

DRIVE_FILE_ID   = "1fPYmvioQvYJjKUMuQgDayf3BnSSEJ7Mp"
HOJAS_EXCLUIDAS = {"CALENDARIO", "RESUMEN_MES", "RESUMEN"}


# ── Crear hoja con formato estándar ──────────────────────────────────────────

def _crear_hoja_proyecto(wb, pid: str, nombre: str):
    """
    Crea pestaña nueva con el esqueleto de columnas estándar.
    Los datos de beneficiarios quedan en blanco — el equipo los completa.
    """
    ws = wb.create_sheet(title=pid)

    # Fila 1 — título
    ws["B1"] = f"{pid} – {nombre}"

    # Fila 2 — meta y SPI placeholder (el script de curvas S lo actualizará)
    ws["B2"] = "Meta: pendiente de completar"
    ws["H2"] = "SPI 0.000"

    # Fila 5 — encabezados de columnas
    headers = [
        None,                  # A (vacía)
        "Nombre Beneficiario", # B
        "Grupo",               # C
        "Capataz",             # D
        "Av. Viv%",           # E
        "Av. Total%",         # F
        "SPI",                 # G
        "Modo",                # H
        "Desp. Real",         # I
        "Jul 2026",            # J  = mes1
        "Ago 2026",            # K  = mes2
        "Sep 2026",            # L  = mes3
        "P10 Dias",            # M
        "P50 Dias",            # N
        "P90 Dias",            # O
    ]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=5, column=col).value = h

    # Fila 6 — marcador de grupo vacío para que el formato sea legible
    ws["B6"] = "  GRUPO 1"

    print(f"  [+] Hoja '{pid}' creada ({nombre})"
          f" — completar beneficiarios y datos P50 manualmente")


# ── Drive: descarga ───────────────────────────────────────────────────────────

def _descargar_excel() -> bytes:
    sys.path.insert(0, str(Path(__file__).parent))
    sys.path.insert(0, str(Path(__file__).parent.parent / "curvas_s"))
    import curvas_cloud_utils as _ccu
    from googleapiclient.discovery import build
    from googleapiclient.http import MediaIoBaseDownload

    creds   = _ccu.get_credentials()
    service = build("drive", "v3", credentials=creds)
    meta    = service.files().get(fileId=DRIVE_FILE_ID, fields="mimeType").execute()

    if meta["mimeType"] == "application/vnd.google-apps.spreadsheet":
        req = service.files().export_media(
            fileId=DRIVE_FILE_ID,
            mimeType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    else:
        req = service.files().get_media(fileId=DRIVE_FILE_ID)

    buf = io.BytesIO()
    dl  = MediaIoBaseDownload(buf, req)
    done = False
    while not done:
        _, done = dl.next_chunk()
    return buf.getvalue()


# ── Drive: subida ─────────────────────────────────────────────────────────────

def _subir_excel(contenido: bytes):
    sys.path.insert(0, str(Path(__file__).parent.parent / "curvas_s"))
    import curvas_cloud_utils as _ccu
    from googleapiclient.discovery import build
    from googleapiclient.http import MediaIoBaseUpload

    creds   = _ccu.get_credentials()
    service = build("drive", "v3", credentials=creds)
    media   = MediaIoBaseUpload(
        io.BytesIO(contenido),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        resumable=True,
    )
    service.files().update(fileId=DRIVE_FILE_ID, media_body=media).execute()
    print("  [Drive] Proyeccion_Despachos_2026.xlsx actualizado en Drive.")


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    print("=" * 60)
    print("Verificacion proyectos — Proyeccion_Despachos_2026.xlsx")
    print("=" * 60)

    print("\nDescargando Excel de Drive...")
    datos = _descargar_excel()

    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(datos))

    hojas_proyecto = set(wb.sheetnames) - HOJAS_EXCLUIDAS
    pids_activos   = {pid for pid, _ in PROYECTOS_ACTIVOS}

    faltantes = [(pid, nom) for pid, nom in PROYECTOS_ACTIVOS
                 if pid not in hojas_proyecto]
    extras    = hojas_proyecto - pids_activos

    print(f"\nProyectos activos ({len(pids_activos)}): {sorted(pids_activos)}")
    print(f"Hojas en Excel   ({len(hojas_proyecto)}): {sorted(hojas_proyecto)}")

    if faltantes:
        print(f"\nProyectos sin hoja ({len(faltantes)}) — creando:")
        for pid, nom in faltantes:
            _crear_hoja_proyecto(wb, pid, nom)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        print("\nSubiendo Excel actualizado a Drive...")
        _subir_excel(buf.read())
        print(f"\nListo — {len(faltantes)} hoja(s) creada(s).")
    else:
        print("\nTodos los proyectos activos tienen hoja en el Excel. Sin cambios.")

    if extras:
        print(f"\nAVISO: Hojas sin proyecto activo correspondiente: {sorted(extras)}")
        print("  (Pueden ser proyectos terminados — verificar si deben eliminarse.)")


if __name__ == "__main__":
    main()
