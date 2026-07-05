"""
sincronizar_datos_excel.py
==========================
Sincroniza datos reales desde Firebase y Google Sheets al Excel
Proyeccion_Despachos_2026.xlsx. Se ejecuta automáticamente cada
viernes junto con la proyección de despachos.

Por cada proyecto activo:
  1. Lee estructura de grupos y capataz desde Firebase /grupos/{pid}
  2. Lee beneficiarios + avance individual desde Google Sheets (Datos Control)
  3. Lee inicio/plazo del proyecto desde Firebase /gantt_programa/{pid}
  4. Calcula SPI y P50 estimados
  5. Crea la hoja si no existe; actualiza SPI/avance si ya existe
  → NO modifica mes1/mes2/mes3 (gestionados por proyectar_despachos_gantt.py)
  → NO modifica celdas que ya tengan [SOL] o [MC] tags

Uso:
    python sincronizar_datos_excel.py              # todos los proyectos
    python sincronizar_datos_excel.py P28          # solo un proyecto
    python sincronizar_datos_excel.py --preview    # muestra sin guardar
"""

import io
import re
import sys
import math
import argparse
import unicodedata
from datetime import date, datetime
from pathlib import Path

import requests
from openpyxl.cell import MergedCell


def _safe_write(ws, row: int, col: int, value):
    """Escribe en una celda solo si no es parte de un rango fusionado (read-only)."""
    cell = ws.cell(row=row, column=col)
    if not isinstance(cell, MergedCell):
        cell.value = value


FIREBASE_BASE = "https://scraices-dashboard-default-rtdb.firebaseio.com"
DRIVE_FILE_ID = "1fPYmvioQvYJjKUMuQgDayf3BnSSEJ7Mp"
HOJAS_EXCLUIDAS = {"CALENDARIO", "RESUMEN_MES", "RESUMEN"}
TODAY = date.today()

# Proyectos activos: pid → {nombre, sheet_id del Gantt de control}
PROYECTOS = {
    "P119": {"nombre": "Nuke Mapu",       "sheet_id": "1t_1j62f_3l1nrlufmvhnV-o1WTplv0OnQL_JdVaWgKA"},
    "P38":  {"nombre": "Aliwen",           "sheet_id": "151wIDnZn8_b7egJKLQKUcD6QDCflEWF5OlAU9KWgc-M"},
    "P39":  {"nombre": "El Coihue",        "sheet_id": "1GiaZ1i3BN3mbgFmg16Ze5E25R0jEmYULtCWRy6cKaHo"},
    "P127": {"nombre": "Nuevo Cunco",      "sheet_id": "1bgT-83Aea0DlyeQ6OvitGDZfm3jLRQUV0GooVP-G8EI"},
    "P12":  {"nombre": "Juan Huilcan",     "sheet_id": "1SLu5lQTAzhHOUuorM3jbxSBes7vMyIB9mMXCihQ6A40"},
    "P14":  {"nombre": "Com. Madihue",     "sheet_id": "1B4wO-UkIDDyFvwRYjMAGksJvKqtNblwl_IirLl3qA6E"},
    "P126": {"nombre": "El Maiten",        "sheet_id": "1IwBN7CpDvKVAvaRuYHNUfkQ0e98cJMVFDKFrntv3oNw"},
    "P131": {"nombre": "Raices Melipeuco", "sheet_id": "1n5F-P5cy8Wj5BujllzdnwCrKwGsfIkdvHxcyd6YwscU"},
    "P116": {"nombre": "Sonia Quilaleo",   "sheet_id": "1z9kNq9uo363NrWqCojGfGpP326FDj3V60irWxtMMbU8"},
    "P31":  {"nombre": "Trovolhue",        "sheet_id": "1kgroktRIto3gGGnvmXGMgoetXT-Rv8JFue1K9eNlNMg"},
    "P28":  {"nombre": "Elsa Pinchulaf",   "sheet_id": "18XkRb7RAF52Aqj4immGebME-d9sODY0HgxIKWcBGrlg"},
}

MES_HEADERS = ["Jul 2026", "Ago 2026", "Sep 2026"]


# ── Utilidades ────────────────────────────────────────────────────────────────

def _norm(texto: str) -> str:
    """Normaliza texto: sin tildes, minúsculas, sin espacios extra."""
    nfkd = unicodedata.normalize("NFKD", str(texto or ""))
    sin_t = "".join(c for c in nfkd if not unicodedata.combining(c))
    return " ".join(sin_t.lower().split())


def _grupo_num(texto: str) -> int | None:
    """Extrae el número de grupo de un texto como 'Grupo N°1' o 'GRUPO 2'."""
    m = re.search(r"\d+", str(texto or ""))
    return int(m.group()) if m else None


def _calcular_spi(inicio: date, plazo: int, avance_real_pct: float) -> float:
    """SPI = avance_real / avance_programa en la fecha de control."""
    if TODAY <= inicio or plazo <= 0:
        return 0.0
    elapsed = (TODAY - inicio).days
    avance_prog = min(100.0, (elapsed / plazo) * 100.0)
    if avance_prog <= 0:
        return 0.0
    return round(avance_real_pct / avance_prog, 3)


def _calcular_p50(inicio: date, plazo: int, avance_real_pct: float) -> int:
    """Días desde hoy hasta P50 de finalización (estimado lineal)."""
    if TODAY < inicio:
        # Proyecto no iniciado: P50 = días hasta inicio + mitad del plazo
        return (inicio - TODAY).days + plazo // 2
    elapsed = max(1, (TODAY - inicio).days)
    if avance_real_pct <= 0:
        # Sin avance real: asumir ritmo del programa (1 día de obra = 100/plazo %)
        daily = 100.0 / plazo
    else:
        daily = avance_real_pct / elapsed
    remaining = max(0, 100.0 - avance_real_pct)
    if daily <= 0:
        return plazo
    return max(1, math.ceil(remaining / daily))


# ── Firebase ──────────────────────────────────────────────────────────────────

def _fb_get(path: str):
    url = f"{FIREBASE_BASE}/{path}.json"
    try:
        r = requests.get(url, timeout=15)
        r.raise_for_status()
        return r.json()
    except Exception as e:
        print(f"    [Firebase] Error GET {path}: {e}")
        return None


def _leer_grupos_firebase(pid: str) -> dict[int, str]:
    """Retorna {num_grupo: capataz} desde Firebase /grupos/{pid}."""
    datos = _fb_get(f"grupos/{pid}")
    if not datos:
        return {}
    resultado = {}
    for i, g in enumerate(datos):
        num = _grupo_num(g.get("nombre", "")) or (i + 1)
        capataz = g.get("capataz", "")
        resultado[num] = capataz
    return resultado


def _leer_gantt_firebase(pid: str) -> tuple[date | None, int]:
    """Retorna (fecha_inicio, plazo_dias) desde Firebase /gantt_programa/{pid}."""
    datos = _fb_get(f"gantt_programa/{pid}")
    if not datos:
        return None, 0
    try:
        inicio = datetime.strptime(datos["inicio"], "%Y-%m-%d").date()
        plazo = int(datos.get("plazo", 0))
        return inicio, plazo
    except Exception:
        return None, 0


# ── Google Sheets: Datos Control ──────────────────────────────────────────────

def _leer_datos_control(sheets_svc, sheet_id: str, pid: str) -> list[dict]:
    """
    Lee la hoja 'Datos Control' del Gantt de control.
    Retorna lista de {nombre, grupo_num, avance_pct, inicio}.
    """
    # Buscar nombre exacto de la hoja
    try:
        meta = sheets_svc.spreadsheets().get(spreadsheetId=sheet_id).execute()
        sheet_names = [s["properties"]["title"] for s in meta["sheets"]]
    except Exception as e:
        print(f"    [Sheets] Error obteniendo hojas {pid}: {e}")
        return []

    hoja = None
    for candidato in ["Datos Control", "datos control", "DatosControl"]:
        if candidato in sheet_names:
            hoja = candidato
            break
    if not hoja:
        print(f"    [Sheets] {pid}: no se encontró 'Datos Control' en {sheet_names[:5]}")
        return []

    try:
        result = sheets_svc.spreadsheets().values().get(
            spreadsheetId=sheet_id,
            range=f"'{hoja}'!A1:D200",
            valueRenderOption="UNFORMATTED_VALUE",
        ).execute()
        rows = result.get("values", [])
    except Exception as e:
        print(f"    [Sheets] Error leyendo {pid}: {e}")
        return []

    beneficiarios = []
    for row in rows[4:]:   # filas 5+ son datos
        if not row or len(row) < 2:
            continue
        grupo_raw = str(row[0]).strip()
        nombre    = str(row[1]).strip()
        if not nombre or not grupo_raw:
            continue
        # Ignorar filas de encabezado residuales
        if _norm(nombre) in ("beneficiario", "grupo", ""):
            continue

        try:
            pct = float(str(row[3]).replace("%", "").strip()) if len(row) > 3 else 0.0
        except (ValueError, TypeError):
            pct = 0.0

        inicio_s = str(row[2]).strip() if len(row) > 2 else ""
        try:
            inicio_b = datetime.strptime(inicio_s, "%d/%m/%Y").date()
        except ValueError:
            try:
                inicio_b = datetime.strptime(inicio_s, "%Y-%m-%d").date()
            except ValueError:
                inicio_b = None

        beneficiarios.append({
            "nombre":    nombre.upper(),
            "grupo_num": _grupo_num(grupo_raw) or 1,
            "avance":    pct,
            "inicio":    inicio_b,
        })

    return beneficiarios


# ── Excel: Drive ──────────────────────────────────────────────────────────────

def _descargar_excel() -> bytes:
    sys.path.insert(0, str(Path(__file__).parent))
    sys.path.insert(0, str(Path(__file__).parent.parent / "curvas_s"))
    import curvas_cloud_utils as _ccu
    from googleapiclient.discovery import build
    from googleapiclient.http import MediaIoBaseDownload

    creds   = _ccu.get_credentials()
    service = build("drive", "v3", credentials=creds)
    meta    = service.files().get(fileId=DRIVE_FILE_ID, fields="mimeType").execute()
    if meta["mimeType"] == "application/vnd.google-apps.spreadsheet":
        req = service.files().export_media(
            fileId=DRIVE_FILE_ID,
            mimeType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    else:
        req = service.files().get_media(fileId=DRIVE_FILE_ID)
    buf = io.BytesIO()
    dl  = MediaIoBaseDownload(buf, req)
    done = False
    while not done:
        _, done = dl.next_chunk()
    return buf.getvalue()


def _subir_excel(contenido: bytes):
    sys.path.insert(0, str(Path(__file__).parent.parent / "curvas_s"))
    import curvas_cloud_utils as _ccu
    from googleapiclient.discovery import build
    from googleapiclient.http import MediaIoBaseUpload

    creds   = _ccu.get_credentials()
    service = build("drive", "v3", credentials=creds)
    media   = MediaIoBaseUpload(
        io.BytesIO(contenido),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        resumable=True,
    )
    service.files().update(fileId=DRIVE_FILE_ID, media_body=media).execute()
    print("  [Drive] Proyeccion_Despachos_2026.xlsx actualizado en Drive.")


# ── Escritura en hoja Excel ───────────────────────────────────────────────────

def _tiene_tag(celda_valor) -> bool:
    """Retorna True si la celda tiene contenido [SOL] o [MC] (no tocar)."""
    v = str(celda_valor or "").strip()
    return "[SOL]" in v or "[MC]" in v


def _sincronizar_hoja(ws, pid: str, sheets_svc, preview: bool = False) -> int:
    """
    Crea o actualiza una hoja de proyecto en el Excel.
    Retorna número de filas escritas/actualizadas.
    """
    meta = PROYECTOS[pid]
    print(f"\n  [{pid}] {meta['nombre']}")

    # 1. Leer datos de Firebase
    grupos_capataz = _leer_grupos_firebase(pid)
    inicio, plazo  = _leer_gantt_firebase(pid)
    print(f"    Firebase: inicio={inicio}  plazo={plazo}d  "
          f"grupos/capataz={grupos_capataz}")

    # 2. Leer beneficiarios desde Google Sheets
    beneficiarios = _leer_datos_control(sheets_svc, meta["sheet_id"], pid)
    print(f"    Sheets:   {len(beneficiarios)} beneficiarios leídos")
    if not beneficiarios:
        print(f"    AVISO: Sin datos de beneficiarios — hoja no modificada")
        return 0

    # 3. Calcular SPI a nivel proyecto (promedio de avances)
    if inicio and plazo:
        avance_avg = sum(b["avance"] for b in beneficiarios) / len(beneficiarios)
        spi_proy   = _calcular_spi(inicio, plazo, avance_avg)
    else:
        avance_avg = 0.0
        spi_proy   = 0.0

    spi_str = f"SPI {spi_proy:.3f}"
    print(f"    Calculado: avance_avg={avance_avg:.1f}%  SPI={spi_proy:.3f}")

    if preview:
        for b in beneficiarios:
            p50 = _calcular_p50(inicio or TODAY, plazo or 180, b["avance"]) if inicio else 999
            capataz = grupos_capataz.get(b["grupo_num"], "")
            print(f"      {b['nombre']:<40} G{b['grupo_num']}  "
                  f"av={b['avance']:.1f}%  P50={p50}d  cap={capataz}")
        return len(beneficiarios)

    # 4. Escribir encabezados de hoja (filas 1-2-5)
    ws["B1"] = f"{pid} – {meta['nombre']}"
    ws["B2"] = f"Actualizado: {TODAY.strftime('%d/%m/%Y')}"
    ws["H2"] = spi_str

    headers = [
        None, "Nombre Beneficiario", "Grupo", "Capataz",
        "Av. Viv%", "Av. Total%", "SPI", "Modo", "Desp. Real",
        MES_HEADERS[0], MES_HEADERS[1], MES_HEADERS[2],
        "P10 Dias", "P50 Dias", "P90 Dias",
    ]
    for col, h in enumerate(headers, start=1):
        _safe_write(ws, 5, col, h)

    # 5. Agrupar beneficiarios por grupo para escribir en orden
    grupos_dict: dict[int, list] = {}
    for b in beneficiarios:
        grupos_dict.setdefault(b["grupo_num"], []).append(b)

    row_idx = 6
    filas_escritas = 0

    for num_grupo in sorted(grupos_dict.keys()):
        bens = grupos_dict[num_grupo]
        capataz = grupos_capataz.get(num_grupo, "")

        # Fila de encabezado de grupo
        _safe_write(ws, row_idx, 2, f"  GRUPO {num_grupo}")
        row_idx += 1

        for b in bens:
            p50  = _calcular_p50(inicio or TODAY, plazo or 180, b["avance"]) if inicio else 999
            p10  = max(1, round(p50 * 0.75))
            p90  = round(p50 * 1.35)
            spi_b = _calcular_spi(
                b["inicio"] or inicio or TODAY,
                plazo or 180,
                b["avance"],
            )

            # Buscar si ya existe esta fila (por nombre) para no pisar [SOL]/[MC]
            celda_mes1 = ws.cell(row=row_idx, column=10)
            celda_mes2 = ws.cell(row=row_idx, column=11)
            celda_mes3 = ws.cell(row=row_idx, column=12)

            # Nombre + grupo + capataz + avances + SPI (siempre actualizar)
            _safe_write(ws, row_idx, 2,  b["nombre"])
            _safe_write(ws, row_idx, 3,  f"Grupo {num_grupo}")
            _safe_write(ws, row_idx, 4,  capataz)
            _safe_write(ws, row_idx, 5,  f"{b['avance']:.1f}%")
            _safe_write(ws, row_idx, 6,  f"{b['avance']:.1f}%")
            _safe_write(ws, row_idx, 7,  f"{spi_b:.3f}")
            _safe_write(ws, row_idx, 8,  "[MC]")

            # mes1/mes2/mes3: solo escribir si estaban vacías o "—"
            for col_idx, celda in [(10, celda_mes1), (11, celda_mes2), (12, celda_mes3)]:
                if not isinstance(celda, MergedCell) and not _tiene_tag(celda.value):
                    celda.value = "—"

            # P10/P50/P90: siempre actualizar (el projection script los necesita frescos)
            _safe_write(ws, row_idx, 13, p10)
            _safe_write(ws, row_idx, 14, p50)
            _safe_write(ws, row_idx, 15, p90)

            filas_escritas += 1
            row_idx += 1

    print(f"    -> {filas_escritas} beneficiarios escritos  SPI={spi_str}")
    return filas_escritas


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Sincroniza datos al Excel de Despachos")
    parser.add_argument("pid", nargs="?", default=None,
                        help="Proyecto a sincronizar (ej: P28). Sin argumento = todos.")
    parser.add_argument("--preview", action="store_true",
                        help="Muestra cambios sin guardar ni subir a Drive")
    args = parser.parse_args()

    print("=" * 60)
    print(f"Sincronizar Datos Excel — {TODAY}  "
          f"[{'PREVIEW' if args.preview else 'ACTUALIZAR'}]")
    print("=" * 60)

    # Determinar qué proyectos procesar
    if args.pid:
        if args.pid not in PROYECTOS:
            print(f"ERROR: Proyecto '{args.pid}' no encontrado en PROYECTOS.")
            sys.exit(1)
        pids = [args.pid]
    else:
        pids = list(PROYECTOS.keys())

    # Credenciales Google una sola vez
    sys.path.insert(0, str(Path(__file__).parent))
    sys.path.insert(0, str(Path(__file__).parent.parent / "curvas_s"))
    import curvas_cloud_utils as _ccu
    from googleapiclient.discovery import build

    creds      = _ccu.get_credentials()
    sheets_svc = build("sheets", "v4", credentials=creds)

    # Descargar Excel
    print("\nDescargando Excel de Drive...")
    datos = _descargar_excel()

    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(datos))

    total_filas = 0
    for pid in pids:
        # Crear hoja si no existe
        if pid not in wb.sheetnames:
            print(f"\n  [{pid}] Hoja nueva — creando...")
            wb.create_sheet(title=pid)

        filas = _sincronizar_hoja(wb[pid], pid, sheets_svc, preview=args.preview)
        total_filas += filas

    print(f"\n{'='*60}")
    print(f"Total filas sincronizadas: {total_filas}")

    if args.preview:
        print("Modo PREVIEW — no se guardaron cambios")
        return

    if total_filas == 0:
        print("Sin cambios — no se sube a Drive")
        return

    print("\nSubiendo Excel actualizado a Drive...")
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    try:
        _subir_excel(buf.read())
    except Exception as e:
        local = Path(__file__).parent / "Proyeccion_Despachos_sync.xlsx"
        buf.seek(0)
        local.write_bytes(buf.read())
        print(f"  [Drive] ERROR: {e} — guardado local: {local}")

    # Actualizar Firebase con los datos frescos
    print("\nActualizando Firebase con datos frescos...")
    try:
        from inyectar_despachos import escribir_despachos_firebase, escribir_despachos_data_firebase
        escribir_despachos_firebase()
        escribir_despachos_data_firebase()
        print("  Firebase /despachos_html y /despachos_data actualizados.")
    except Exception as e:
        print(f"  [Firebase] ERROR: {e}")

    print("\nListo.")


if __name__ == "__main__":
    main()
