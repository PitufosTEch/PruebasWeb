"""
proyectar_despachos_gantt.py
============================
Actualiza las proyecciones de despacho en Proyeccion_Despachos_2026.xlsx
respetando la secuencia del programa de obra (Gantt de control).

Metodología:
  - Entradas [SOL] = solicitudes confirmadas → NO se mueven (quedan en su mes)
  - Entradas [MC]  = proyecciones Monte Carlo → se re-proyectan con la secuencia Gantt
  - SPI_efectivo = min(SPI_real, SPI_objetivo=1.15) → cap: nunca más agresivo que 1.15
  - av_viv por beneficiario leído desde Firebase /avance_benef/{pid}
  - Ruta crítica desde config/etapas_config.json (duración y dependencias por etapa)
  - Pipeline de 10 días entre beneficiarios consecutivos del grupo (ajustado por SPI)
  - Lógica de secuenciamiento:
      · Beneficiario activo (av_viv > 0): ben_start = hoy - (av_viv% × CP/SPI)
      · Beneficiario no iniciado (av_viv = 0): ben_start = prev_start + pipeline/SPI
  - Cada etapa [MC] se proyecta a su fecha específica según dependencias de etapas_config
  - Etapas fuera del rango jul-sep 2026 se eliminan de la vista (→ "—")

Uso:
    python proyectar_despachos_gantt.py [--spi 1.15] [--preview]
    --spi N     Factor SPI objetivo (default: 1.15)
    --preview   Muestra cambios sin guardar ni subir a Drive
"""

import io
import re
import sys
import json
import math
import unicodedata
import argparse
import requests
from calendar import monthrange
from datetime import date, timedelta
from pathlib import Path

SPI_OBJETIVO_DEFAULT = 1.15
PIPELINE_DIAS        = 10      # días de pipeline entre beneficiarios consecutivos

TODAY = date.today()

# ── Ventana dinámica: mes actual + 2 siguientes ───────────────────────────────

def _add_months(d: date, n: int) -> date:
    m = d.month - 1 + n
    return date(d.year + m // 12, m % 12 + 1, 1)

def _ultimo_dia(d: date) -> date:
    return date(d.year, d.month, monthrange(d.year, d.month)[1])

_m0  = TODAY.replace(day=1)
MES1 = (_m0,                 _ultimo_dia(_m0))
MES2 = (_add_months(_m0, 1), _ultimo_dia(_add_months(_m0, 1)))
MES3 = (_add_months(_m0, 2), _ultimo_dia(_add_months(_m0, 2)))

_MESES_NOMBRE = {1:"Ene",2:"Feb",3:"Mar",4:"Abr",5:"May",6:"Jun",
                 7:"Jul",8:"Ago",9:"Sep",10:"Oct",11:"Nov",12:"Dic"}
_MESES_NUM    = {v: k for k, v in _MESES_NOMBRE.items()}

def _mes_str(d: date) -> str:
    return f"{_MESES_NOMBRE[d.month]} {d.year}"

def _parse_mes_header(h) -> date | None:
    """Parsea 'Ago 2026' → date(2026, 8, 1)."""
    m = re.match(r"([A-Za-z]{3})\s+(\d{4})", str(h or "").strip())
    if not m:
        return None
    num = _MESES_NUM.get(m.group(1).capitalize())
    return date(int(m.group(2)), num, 1) if num else None

DRIVE_FILE_ID    = "1fPYmvioQvYJjKUMuQgDayf3BnSSEJ7Mp"
HOJAS_EXCLUIDAS  = {"CALENDARIO", "RESUMEN_MES", "RESUMEN"}
FIREBASE_ROOT    = "https://scraices-dashboard-default-rtdb.firebaseio.com"


# ── Utilidades de fecha ───────────────────────────────────────────────────────

def _date_to_mes(d: date) -> int:
    if MES1[0] <= d <= MES1[1]: return 1
    if MES2[0] <= d <= MES2[1]: return 2
    if MES3[0] <= d <= MES3[1]: return 3
    return 0


# ── Parseo/formato de celdas ─────────────────────────────────────────────────

def _parse_etapas_celda(texto) -> list[dict]:
    texto = str(texto or "").strip()
    if texto in ("—", "-", "", "None"):
        return []
    resultado = []
    for parte in texto.split(","):
        parte = parte.strip()
        m = re.match(r"^\[(MC|SOL)\]\s*(.+)$", parte)
        if m:
            resultado.append({"tag": m.group(1), "nombre": m.group(2).strip()})
        elif parte:
            resultado.append({"tag": "MC", "nombre": parte})
    return resultado


def _formatear_celda(etapas: list[dict]) -> str:
    if not etapas:
        return "—"
    return ", ".join(f"[{e['tag']}] {e['nombre']}" for e in etapas)


def _parse_pct(val) -> float:
    try:
        return float(str(val or "0").replace("%", "").strip())
    except (ValueError, TypeError):
        return 0.0


def _codigo_de_etapa(nombre: str) -> str:
    """Extrae el código numérico del inicio del nombre: '01 Fundaciones' → '01'."""
    parte = str(nombre or "").strip().split()[0] if nombre else ""
    return re.sub(r"[^0-9]", "", parte)[:2]


# ── Normalización de claves Firebase ────────────────────────────────────────

def _normalizar_clave(nombre: str) -> str:
    nfkd = unicodedata.normalize("NFKD", str(nombre))
    ascii_s = "".join(c for c in nfkd if not unicodedata.combining(c))
    return re.sub(r"\s+", "_", ascii_s.upper().strip())


# ── Firebase: av_viv por beneficiario ───────────────────────────────────────

def _fetch_avance_benef(pid: str) -> dict:
    try:
        resp = requests.get(f"{FIREBASE_ROOT}/avance_benef/{pid}.json", timeout=15)
        if resp.status_code == 200:
            data = resp.json()
            return data if isinstance(data, dict) else {}
    except Exception:
        pass
    return {}


# ── Matching robusto de nombres ──────────────────────────────────────────────

def _match_nombre(a: str, b: str) -> bool:
    """
    Compara dos nombres tolerando: tildes, orden de palabras, mayúsculas.
    Pasa 3 filtros en cascada:
      1. Token-set exacto (después de normalizar)
      2. Jaccard >= 0.75
      3. Coincidencia de las 2 palabras más cortas (apellidos en conv. chilena)
    """
    def _tokens(s: str) -> set:
        nfkd = unicodedata.normalize("NFKD", str(s))
        ascii_s = "".join(c for c in nfkd if not unicodedata.combining(c))
        return set(re.sub(r"[^A-Z0-9]", " ", ascii_s.upper()).split())

    ta, tb = _tokens(a), _tokens(b)
    if not ta or not tb:
        return False
    # Paso 1: conjuntos idénticos
    if ta == tb:
        return True
    # Paso 2: Jaccard >= 0.75
    inter = ta & tb
    union = ta | tb
    if union and len(inter) / len(union) >= 0.75:
        return True
    # Paso 3: los 2 tokens ordenados alfabéticamente coinciden
    # (apellidos chilenos tienden a ser únicos; el sort neutraliza el orden)
    sa, sb = sorted(ta), sorted(tb)
    if len(sa) >= 2 and len(sb) >= 2 and sa[:2] == sb[:2]:
        return True
    return False


def _fetch_despachados(pid: str) -> list:
    """Lee /sol_despachados/{pid} → [{nombre, etapa}] desde Firebase."""
    try:
        resp = requests.get(f"{FIREBASE_ROOT}/sol_despachados/{pid}.json", timeout=15)
        if resp.status_code == 200:
            data = resp.json()
            return data if isinstance(data, list) else []
    except Exception:
        pass
    return []


# ── etapas_config.json ────────────────────────────────────────────────────────

def _cargar_etapas_cfg() -> dict:
    path = Path(__file__).parent.parent / "config" / "etapas_config.json"
    with open(path, encoding="utf-8") as f:
        return json.load(f)


def _nombres_etapas(cfg: dict) -> dict[str, str]:
    """Retorna {codigo: 'NN Nombre'} para generar [MC] en filas vacías."""
    result = {}
    for key, etapa in cfg["etapas"].items():
        codigo = etapa.get("codigo", "")
        if not codigo:
            continue
        partes = key.split("_")
        nombre = " ".join(p.capitalize() for p in partes[1:]) if len(partes) > 1 else key.capitalize()
        result[codigo] = f"{codigo} {nombre}"
    return result


def _cp_dias(cfg: dict) -> float:
    """Suma de duraciones de la secuencia principal."""
    return sum(
        cfg["etapas"].get(k, {}).get("duracion", 0)
        for k in cfg.get("secuencia_principal", [])
    )


# ── Schedule de etapas desde una fecha de inicio ────────────────────────────

def _build_schedule(cfg: dict, ben_start: date, spi_ef: float) -> dict[str, date]:
    """
    Retorna {codigo: fecha_dispatch} para cada etapa definida en cfg.
    - Etapas 'flexible' o 'desde_inicio_dependencia' se despachan el mismo día que su dependencia.
    - Entre grupos de despacho consecutivos: mínimo max(duracion_dep/SPI, 7) días.
    """
    etapas = cfg["etapas"]
    cache: dict[str, date] = {}

    def _dispatch(key: str) -> date:
        if key in cache:
            return cache[key]
        etapa = etapas.get(key, {})
        dep   = etapa.get("dependencia")

        if dep is None:
            d = ben_start
        elif etapa.get("flexible") or etapa.get("desde_inicio_dependencia"):
            # Despachar junto con la dependencia (mismo día de despacho)
            d = _dispatch(dep)
        else:
            dep_date = _dispatch(dep)
            dur_dep  = etapas.get(dep, {}).get("duracion", 0) / spi_ef
            d = dep_date + timedelta(days=max(dur_dep, 7))

        cache[key] = d
        return d

    for key in etapas:
        _dispatch(key)

    return {etapas[k]["codigo"]: cache[k] for k in etapas}


# ── Proyección de una fila (beneficiario/grupo) ──────────────────────────────

def _proyectar_fila(
    mes1_val, mes2_val, mes3_val,
    ben_start: date,
    schedule: dict[str, date],
    cp_dias_tot: float,
    spi_ef: float,
    etapas_desp: frozenset = frozenset(),
    mes_offset: int = 0,
    nombres_etapas: dict = {},
) -> tuple[str, str, str, int | None]:
    """
    Distribuye las etapas [MC] de una fila según el schedule del Gantt.
    - [SOL] se desplazan mes_offset meses hacia adelante si la ventana avanzó.
    - Si no hay [MC] existentes, los genera desde el schedule completo.
    - Devuelve (nuevo_mes1, nuevo_mes2, nuevo_mes3, p50_dias).
    """
    etapas_orig = {
        1: _parse_etapas_celda(mes1_val),
        2: _parse_etapas_celda(mes2_val),
        3: _parse_etapas_celda(mes3_val),
    }

    # [SOL]: desplazar si la ventana avanzó (mes pasado → MES1 actual)
    sol_por_mes: dict[int, list] = {1: [], 2: [], 3: []}
    mc_lista: list[dict] = []
    for m in (1, 2, 3):
        for e in etapas_orig[m]:
            if e["tag"] == "SOL":
                nuevo_m = max(1, m - mes_offset)
                if 1 <= nuevo_m <= 3:
                    sol_por_mes[nuevo_m].append(e)
            else:
                mc_lista.append(e)

    nuevas_mc: dict[int, list] = {1: [], 2: [], 3: []}

    if mc_lista:
        n = len(mc_lista)
        for k, etapa in enumerate(mc_lista, start=1):
            codigo = _codigo_de_etapa(etapa["nombre"])

            # Excluir etapas confirmadas como Despachado en AppSheet
            if codigo and codigo in etapas_desp:
                continue

            if codigo in schedule:
                fecha = schedule[codigo]
            else:
                # Fallback lineal para etapas no en config (RC, Quincallería, etc.)
                frac = k / n
                fecha = ben_start + timedelta(days=frac * cp_dias_tot / spi_ef)

            if fecha < TODAY:
                continue

            mes = _date_to_mes(fecha)
            if mes:
                nuevas_mc[mes].append(etapa)

    elif nombres_etapas:
        # Fila vacía: generar [MC] desde el schedule completo
        for codigo, fecha in sorted(schedule.items(), key=lambda x: x[1]):
            if codigo in etapas_desp:
                continue
            if fecha < TODAY:
                continue
            mes = _date_to_mes(fecha)
            if mes and codigo in nombres_etapas:
                nuevas_mc[mes].append({"tag": "MC", "nombre": nombres_etapas[codigo]})

    else:
        # Sin [MC] y sin config: devolver solo [SOL] desplazados
        completion = ben_start + timedelta(days=cp_dias_tot / spi_ef)
        p50 = max(1, (completion - TODAY).days)
        return (
            _formatear_celda(sol_por_mes[1]),
            _formatear_celda(sol_por_mes[2]),
            _formatear_celda(sol_por_mes[3]),
            p50,
        )

    nuevo = {m: sol_por_mes[m] + nuevas_mc[m] for m in (1, 2, 3)}

    # P50: días desde hoy hasta que el beneficiario termine su ruta crítica
    completion = ben_start + timedelta(days=cp_dias_tot / spi_ef)
    p50 = max(1, (completion - TODAY).days)

    return (
        _formatear_celda(nuevo[1]),
        _formatear_celda(nuevo[2]),
        _formatear_celda(nuevo[3]),
        p50,
    )


# ── Limpieza de [MC] stale para beneficiarios terminados ────────────────────

def _limpiar_mc_stale(ws, fila: dict, preview: bool) -> bool:
    """
    Para un beneficiario terminado (av_viv=100%): si alguna celda mes1/mes2/mes3
    contiene items [MC], los borra (los [SOL] se conservan).
    Retorna True si hubo cambio.
    """
    cambio = False
    for key in ("m1c", "m2c", "m3c"):
        cell = fila[key]
        etapas = _parse_etapas_celda(cell.value)
        solo_sol = [e for e in etapas if e["tag"] == "SOL"]
        tiene_mc = any(e["tag"] == "MC" for e in etapas)
        if tiene_mc:
            nuevo = _formatear_celda(solo_sol)
            if not preview:
                cell.value = nuevo
            cambio = True
    return cambio


# ── Procesamiento de cada hoja ───────────────────────────────────────────────

def _procesar_hoja(ws, spi_objetivo: float, cfg: dict, nombres_eta: dict, preview: bool = False) -> int:
    pid = ws.title

    # SPI
    h2 = str(ws["H2"].value or "").replace("SPI", "").strip()
    try:
        spi_real = float(h2)
        if spi_real < 0.1:
            spi_real = 1.0
    except ValueError:
        spi_real = 1.0

    spi_ef = min(spi_real, spi_objetivo)
    factor = spi_ef / spi_objetivo
    print(f"  SPI real={spi_real:.4f}  →  efectivo={spi_ef:.4f}  objetivo={spi_objetivo}  "
          f"(factor: ×{factor:.3f})  pipeline={PIPELINE_DIAS/spi_ef:.1f}d")

    cp = _cp_dias(cfg)
    pipeline_aj = PIPELINE_DIAS / spi_ef

    # Calcular desplazamiento de ventana respecto al Excel
    old_m1 = _parse_mes_header(ws.cell(5, 10).value)
    if old_m1:
        mes_offset = (MES1[0].year - old_m1.year) * 12 + (MES1[0].month - old_m1.month)
        mes_offset = max(0, mes_offset)
    else:
        mes_offset = 0

    # Actualizar encabezados de columna si la ventana avanzó
    if mes_offset > 0:
        if not preview:
            ws.cell(5, 10).value = _mes_str(MES1[0])
            ws.cell(5, 11).value = _mes_str(MES2[0])
            ws.cell(5, 12).value = _mes_str(MES3[0])
        print(f"  [Ventana] desplazada {mes_offset} mes(es) → "
              f"{_mes_str(MES1[0])} · {_mes_str(MES2[0])} · {_mes_str(MES3[0])}")

    # av_viv y despachados desde Firebase
    avance_benef    = _fetch_avance_benef(pid)
    despachados_pid = _fetch_despachados(pid)   # [{nombre, etapa}] últimos 90 días

    # Leer TODAS las filas de beneficiarios (no solo las que tienen datos en mes)
    filas = []
    for row_idx in range(6, ws.max_row + 1):
        nombre = str(ws.cell(row_idx, 2).value or "").strip()
        if not nombre:
            continue
        # Saltar filas de encabezado de grupo (GRUPO 1, GRUPO 2, GRUPO X, etc.)
        if re.match(r"^\s*GRUPO\b", nombre, re.IGNORECASE):
            continue

        m1c = ws.cell(row_idx, 10)
        m2c = ws.cell(row_idx, 11)
        m3c = ws.cell(row_idx, 12)

        # Saltar filas con celdas combinadas (separadores de sección)
        if any(type(c).__name__ == "MergedCell" for c in (m1c, m2c, m3c)):
            continue

        # av_viv: Firebase primero, luego col E del Excel
        fb_key = _normalizar_clave(nombre)
        fb_val = avance_benef.get(fb_key)
        if fb_val is not None and isinstance(fb_val, (int, float)):
            av = float(fb_val)
        else:
            av = _parse_pct(ws.cell(row_idx, 5).value)

        filas.append({
            "row_idx": row_idx,
            "nombre":  nombre,
            "av_viv":  av,
            "m1c": m1c, "m2c": m2c, "m3c": m3c,
        })

    if not filas:
        return 0

    # Calcular ben_start para cada fila en secuencia
    # Lead = fila con mayor av_viv activa (0 < av < 100)
    activas = [(i, f) for i, f in enumerate(filas) if 0 < f["av_viv"] < 100]
    if activas:
        lead_i, lead = max(activas, key=lambda x: x[1]["av_viv"])
        days_done = (lead["av_viv"] / 100.0) * cp / spi_ef
        lead_start = TODAY - timedelta(days=days_done)
    else:
        # Sin activas: si todas en 0% → aún no empiezan, desde hoy
        lead_i = 0
        lead_start = TODAY

    # Asignar ben_start a cada fila
    prev_start = None
    for i, fila in enumerate(filas):
        if fila["av_viv"] >= 100.0:
            fila["ben_start"] = None  # terminado
        elif fila["av_viv"] > 0:
            days_done = (fila["av_viv"] / 100.0) * cp / spi_ef
            fila["ben_start"] = TODAY - timedelta(days=days_done)
            prev_start = fila["ben_start"]
        else:
            # No iniciado: pipeline desde la fila previa activa/no-iniciada
            if prev_start is not None:
                fila["ben_start"] = prev_start + timedelta(days=pipeline_aj)
            else:
                # No hay previo activo → offset desde el lead
                fila["ben_start"] = lead_start + timedelta(days=(i - lead_i) * pipeline_aj)
            prev_start = fila["ben_start"]

    # Proyectar cada fila
    modificadas = 0
    for fila in filas:
        if fila.get("ben_start") is None:
            # Terminado (av_viv >= 100%): limpiar cualquier [MC] stale que quede en el Excel
            if _limpiar_mc_stale(ws, fila, preview):
                modificadas += 1
                print(f"    {fila['nombre'][:35]:<35}  av=100% — limpiando [MC] stale")
            continue

        ben_start = fila["ben_start"]
        schedule  = _build_schedule(cfg, ben_start, spi_ef)

        # Códigos de etapa ya despachadas para este beneficiario (AppSheet)
        etapas_desp = frozenset(
            _codigo_de_etapa(d["etapa"])
            for d in despachados_pid
            if _match_nombre(fila["nombre"], d["nombre"])
            and _codigo_de_etapa(d["etapa"])
        )

        n_mes1, n_mes2, n_mes3, p50 = _proyectar_fila(
            fila["m1c"].value, fila["m2c"].value, fila["m3c"].value,
            ben_start, schedule, cp, spi_ef,
            etapas_desp=etapas_desp,
            mes_offset=mes_offset,
            nombres_etapas=nombres_eta,
        )

        cambio = (
            str(fila["m1c"].value or "") != n_mes1 or
            str(fila["m2c"].value or "") != n_mes2 or
            str(fila["m3c"].value or "") != n_mes3
        )

        if cambio:
            modificadas += 1
            start_str = ben_start.strftime("%d-%b")
            print(f"    {fila['nombre'][:35]:<35}  av={fila['av_viv']:.0f}%  "
                  f"start:{start_str}  p50:{p50}d")

            if not preview:
                fila["m1c"].value = n_mes1
                fila["m2c"].value = n_mes2
                fila["m3c"].value = n_mes3
                if p50 is not None:
                    ri = fila["row_idx"]
                    ws.cell(ri, 14).value = p50
                    ws.cell(ri, 13).value = max(1, round(p50 * 0.75))
                    ws.cell(ri, 15).value = round(p50 * 1.35)

    return modificadas


# ── Descarga y subida Drive ───────────────────────────────────────────────────

def _descargar_excel() -> bytes:
    sys.path.insert(0, str(Path(__file__).parent))
    sys.path.insert(0, str(Path(__file__).parent.parent / "curvas_s"))
    import curvas_cloud_utils as _ccu
    from googleapiclient.discovery import build
    from googleapiclient.http import MediaIoBaseDownload

    creds   = _ccu.get_credentials()
    service = build("drive", "v3", credentials=creds)
    meta    = service.files().get(fileId=DRIVE_FILE_ID, fields="mimeType").execute()
    if meta["mimeType"] == "application/vnd.google-apps.spreadsheet":
        req = service.files().export_media(
            fileId=DRIVE_FILE_ID,
            mimeType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    else:
        req = service.files().get_media(fileId=DRIVE_FILE_ID)

    buf = io.BytesIO()
    dl  = MediaIoBaseDownload(buf, req)
    done = False
    while not done:
        _, done = dl.next_chunk()
    return buf.getvalue()


def _subir_excel(contenido: bytes) -> bool:
    sys.path.insert(0, str(Path(__file__).parent.parent / "curvas_s"))
    import curvas_cloud_utils as _ccu
    from googleapiclient.discovery import build
    from googleapiclient.http import MediaIoBaseUpload

    creds   = _ccu.get_credentials()
    service = build("drive", "v3", credentials=creds)
    media   = MediaIoBaseUpload(
        io.BytesIO(contenido),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        resumable=True,
    )
    service.files().update(fileId=DRIVE_FILE_ID, media_body=media).execute()
    print("  [Drive] Proyeccion_Despachos_2026.xlsx actualizado en Drive ✓")
    return True


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Re-proyecta despachos con secuencia Gantt")
    parser.add_argument("--spi", type=float, default=SPI_OBJETIVO_DEFAULT)
    parser.add_argument("--preview", action="store_true")
    args = parser.parse_args()

    spi_obj = args.spi
    preview  = args.preview

    print(f"{'='*60}")
    print(f"Proyectar Despachos Gantt — SPI objetivo: {spi_obj}  Pipeline: {PIPELINE_DIAS}d")
    print(f"Fecha base: {TODAY}   Modo: {'PREVIEW' if preview else 'ACTUALIZAR'}")
    print(f"{'='*60}\n")

    cfg         = _cargar_etapas_cfg()
    cp          = _cp_dias(cfg)
    nombres_eta = _nombres_etapas(cfg)
    print(f"Ruta crítica: {cp:.0f} días  ({' → '.join(cfg.get('secuencia_principal', []))})")
    print(f"Ventana:      {_mes_str(MES1[0])} · {_mes_str(MES2[0])} · {_mes_str(MES3[0])}\n")

    print("► Descargando Excel de Drive...")
    try:
        datos = _descargar_excel()
    except Exception as e:
        print(f"  ERROR descargando: {e}")
        sys.exit(1)

    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(datos))
    project_sheets = sorted(s for s in wb.sheetnames if s not in HOJAS_EXCLUIDAS)

    total = 0
    for pid in project_sheets:
        print(f"\n► {pid}")
        mod = _procesar_hoja(wb[pid], spi_obj, cfg, nombres_eta, preview=preview)
        print(f"  → {mod} filas actualizadas")
        total += mod

    print(f"\n{'='*60}")
    print(f"Total filas modificadas: {total}")

    if preview:
        print("Modo PREVIEW — no se guardaron cambios")
        return

    if total == 0:
        print("Sin cambios — no se sube a Drive")
        return

    print("\n► Subiendo Excel actualizado a Drive...")
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    try:
        _subir_excel(buf.read())
    except Exception as e:
        local = Path(__file__).parent / "Proyeccion_Despachos_SPI115.xlsx"
        buf.seek(0)
        local.write_bytes(buf.read())
        print(f"  [Drive] ERROR: {e}")
        print(f"  Guardado local: {local}")

    print("\n► Actualizando Firebase con los nuevos datos...")
    try:
        from inyectar_despachos import (
            escribir_despachos_firebase,
            escribir_despachos_data_firebase,
        )
        escribir_despachos_firebase()
        escribir_despachos_data_firebase()
        print("  Firebase /despachos_html y /despachos_data actualizados ✓")
    except Exception as e:
        print(f"  [Firebase] ERROR: {e}")

    print("\nListo.")


if __name__ == "__main__":
    main()
