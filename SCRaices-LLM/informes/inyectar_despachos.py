"""
inyectar_despachos.py
=====================
Descarga Proyeccion_Despachos_2026.xlsx desde Google Drive, lee las pestañas
por proyecto (P116, P119, ...) y agrega un tab "Despachos" al HTML de
Informe_Adquisiciones, integrado con el selector de obra existente.

Flujo:
  1. Descarga el Excel desde Drive (credenciales OAuth existentes)
  2. Por cada proyecto, parsea beneficiarios + etapas planificadas por mes
  3. Inyecta en el HTML:
     a) Botón de tab "Despachos"
     b) <div class="section" id="sec-despachos">
     c) <script> que extiende el objeto secciones[] con los datos de despacho
"""

import io
import json
import sys
from pathlib import Path

DRIVE_FILE_ID = "1fPYmvioQvYJjKUMuQgDayf3BnSSEJ7Mp"

# Mapa idx HTML → pestaña Excel (según selector del Reporte Adquisiciones)
IDX_TO_PID = {
    0: "P131",
    1: "P126",
    2: "P127",
    3: "P39",
    4: "P14",
    5: "P116",
    6: "P12",
    7: "P38",
    8: "P119",
    9: "P31",
}


# ── Descarga ─────────────────────────────────────────────────────────────────

def _descargar_excel() -> bytes:
    sys.path.insert(0, str(Path(__file__).parent))
    import curvas_cloud_utils as _ccu
    from googleapiclient.discovery import build
    from googleapiclient.http import MediaIoBaseDownload

    creds   = _ccu.get_credentials()
    service = build("drive", "v3", credentials=creds)
    meta    = service.files().get(fileId=DRIVE_FILE_ID, fields="mimeType").execute()

    if meta["mimeType"] == "application/vnd.google-apps.spreadsheet":
        req = service.files().export_media(
            fileId=DRIVE_FILE_ID,
            mimeType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    else:
        req = service.files().get_media(fileId=DRIVE_FILE_ID)

    buf = io.BytesIO()
    dl  = MediaIoBaseDownload(buf, req)
    done = False
    while not done:
        _, done = dl.next_chunk()
    return buf.getvalue()


# ── Parser por proyecto ───────────────────────────────────────────────────────

def _leer_proyecto(ws) -> dict:
    """
    Parsea una pestaña de proyecto y retorna:
    {titulo, meta, spi_str, meses:[str], beneficiarios:[{...}]}
    """
    rows = list(ws.iter_rows(values_only=True))
    titulo   = str(rows[0][1] or "").strip() if rows else ""
    meta_raw = str(rows[1][1] or "").strip() if len(rows) > 1 else ""
    spi_str  = str(rows[1][7] or "").strip() if len(rows) > 1 else ""

    # Encabezados en fila 5 (idx 4)
    if len(rows) > 4:
        hdrs  = [str(v or "").replace("\n", " ").strip() for v in rows[4][1:16]]
        meses = [h for h in hdrs[8:11] if h and h not in ("P10 Días", "P50 Días", "P90 Días")]
    else:
        meses = []

    beneficiarios = []
    for row in rows[5:]:
        if all(v is None for v in row):
            continue
        b = str(row[1] or "").strip()
        c = row[2]
        if not b:
            continue
        # Fila de grupo: col B tiene texto, col C en blanco
        if c is None:
            continue

        def _v(r, i):
            return str(r[i] or "").strip() if i < len(r) else ""

        beneficiarios.append({
            "nombre":    b,
            "grupo":     _v(row, 2),
            "capataz":   _v(row, 3),
            "av_viv":    _v(row, 4),
            "av_total":  _v(row, 5),
            "spi":       _v(row, 6),
            "modo":      _v(row, 7),
            "desp_real": _v(row, 8),
            "mes1":      _v(row, 9),
            "mes2":      _v(row, 10),
            "mes3":      _v(row, 11),
            "p50":       _v(row, 13),
        })

    return {
        "titulo":        titulo,
        "meta":          meta_raw,
        "spi_str":       spi_str,
        "meses":         meses,
        "beneficiarios": beneficiarios,
    }


def _leer_todos_proyectos(data: bytes) -> dict:
    """Retorna {pid: datos_proyecto} para todas las pestañas de proyecto."""
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    resultado = {}
    for pid in IDX_TO_PID.values():
        if pid in wb.sheetnames:
            resultado[pid] = _leer_proyecto(wb[pid])
        else:
            resultado[pid] = None
    return resultado


# ── Generador HTML por proyecto ───────────────────────────────────────────────

def _spi_color(spi_str: str) -> str:
    try:
        v = float(spi_str.replace("SPI", "").strip())
        if v >= 0.95:
            return "#16a34a"
        if v >= 0.75:
            return "#d97706"
        return "#dc2626"
    except Exception:
        return "#6b7280"


def _avance_color(av_str: str) -> str:
    try:
        v = float(av_str)
        if v >= 80:
            return "#16a34a"
        if v >= 40:
            return "#d97706"
        return "#dc2626"
    except Exception:
        return "#374151"


def _parsear_etapas(texto: str) -> list:
    """'[MC] 05 Ap. Exterior, [SOL] 02 1era...' → [('05','Ap. Exterior'), ...]"""
    import re
    if not texto or texto == "—":
        return []
    resultado = []
    for parte in texto.split(","):
        parte = re.sub(r"^\[(MC|SOL)\]\s*", "", parte.strip()).strip()
        m = re.match(r"^(\d+)\s+(.+)$", parte)
        if m:
            resultado.append((m.group(1).zfill(2), m.group(2).strip()))
    return resultado


def _resumen_por_etapa(bens: list, meses: list) -> str:
    """Genera tabla HTML de resumen: etapa × mes → cantidad de beneficiarios."""
    from collections import defaultdict
    conteo = defaultdict(lambda: [0, 0, 0])  # (cod, nombre) → [mes1, mes2, mes3]
    nombres_etapa = {}

    for b in bens:
        for i, key in enumerate(["mes1", "mes2", "mes3"]):
            for cod, nombre in _parsear_etapas(b[key]):
                conteo[(cod, nombre)][i] += 1
                nombres_etapa[cod] = nombre

    if not conteo:
        return ""

    filas = sorted(conteo.items(), key=lambda x: x[0][0])
    totales = [0, 0, 0]

    ths = "".join(
        f'<th style="padding:7px 10px;background:#0f172a;color:#94a3b8;'
        f'font-size:10px;text-align:center;white-space:nowrap;">{m}</th>'
        for m in (meses or ["Mes 1", "Mes 2", "Mes 3"])
    )

    filas_html = ""
    for i, ((cod, nombre), cnts) in enumerate(filas):
        bg = "#ffffff" if i % 2 == 0 else "#f8fafc"
        for j in range(3):
            totales[j] += cnts[j]

        def _cel(v):
            if v == 0:
                return (f'<td style="padding:6px 10px;text-align:center;'
                        f'font-size:12px;color:#d1d5db;border-bottom:1px solid #e2e8f0;">—</td>')
            return (f'<td style="padding:6px 10px;text-align:center;font-size:13px;'
                    f'font-weight:700;color:#1e293b;border-bottom:1px solid #e2e8f0;">{v}</td>')

        filas_html += (
            f'<tr style="background:{bg};">'
            f'<td style="padding:6px 12px;font-size:11px;color:#64748b;'
            f'text-align:center;border-bottom:1px solid #e2e8f0;">{cod}</td>'
            f'<td style="padding:6px 12px;font-size:12px;color:#1e293b;font-weight:500;'
            f'border-bottom:1px solid #e2e8f0;">{nombre}</td>'
            + "".join(_cel(c) for c in cnts) +
            f'<td style="padding:6px 10px;text-align:center;font-size:13px;'
            f'font-weight:700;color:#7c3aed;border-bottom:1px solid #e2e8f0;">'
            f'{sum(cnts) or "—"}</td>'
            f'</tr>'
        )

    tot_celdas = "".join(
        f'<td style="padding:7px 10px;text-align:center;font-size:13px;'
        f'font-weight:700;color:#ffffff;background:#0f172a;">{t if t else "—"}</td>'
        for t in totales
    )
    fila_total = (
        f'<tr>'
        f'<td style="padding:7px 12px;background:#0f172a;"></td>'
        f'<td style="padding:7px 12px;font-size:11px;font-weight:700;color:#94a3b8;'
        f'background:#0f172a;">TOTAL DESPACHOS</td>'
        + tot_celdas +
        f'<td style="padding:7px 10px;text-align:center;font-size:13px;'
        f'font-weight:700;color:#a78bfa;background:#0f172a;">{sum(totales)}</td>'
        f'</tr>'
    )

    return (
        '<div style="margin-top:24px;">'
        '<div style="background:#0f172a;border-radius:8px 8px 0 0;padding:10px 16px;">'
        '<div style="font-size:11px;font-weight:700;color:#e2e8f0;letter-spacing:.05em;'
        'text-transform:uppercase;">Resumen por etapa · Beneficiarios a despachar</div>'
        '</div>'
        '<div style="overflow-x:auto;border:1px solid #e2e8f0;border-radius:0 0 8px 8px;">'
        '<table style="border-collapse:collapse;width:100%;min-width:500px;">'
        '<thead><tr>'
        '<th style="padding:7px 12px;background:#1e293b;color:#94a3b8;font-size:10px;'
        'text-align:center;width:48px;">CÓD.</th>'
        '<th style="padding:7px 12px;background:#1e293b;color:#94a3b8;font-size:10px;'
        'text-align:left;">ETAPA</th>'
        + ths +
        '<th style="padding:7px 10px;background:#0f172a;color:#94a3b8;'
        'font-size:10px;text-align:center;">TOTAL</th>'
        '</tr></thead>'
        f'<tbody>{filas_html}</tbody>'
        f'<tfoot>{fila_total}</tfoot>'
        '</table></div></div>'
    )


def _formatear_etapas(texto: str) -> str:
    """Convierte '[MC] 05 Ap. Exterior, [SOL] 02 1era...' en badges HTML."""
    if not texto or texto == "—":
        return '<span style="color:#d1d5db;">—</span>'
    partes = [p.strip() for p in texto.split(",") if p.strip()]
    html = ""
    for p in partes:
        if p.startswith("[SOL]"):
            color = "#0f766e"
            bg    = "#ccfbf1"
            label = p[5:].strip()
        elif p.startswith("[MC]"):
            color = "#92400e"
            bg    = "#fef3c7"
            label = p[4:].strip()
        else:
            color = "#374151"
            bg    = "#f3f4f6"
            label = p
        html += (
            f'<span style="display:inline-block;margin:1px 2px;padding:1px 5px;'
            f'border-radius:4px;background:{bg};color:{color};font-size:10px;'
            f'white-space:nowrap;">{label}</span>'
        )
    return html


def _generar_html_proyecto(datos: dict) -> str:
    if not datos or not datos.get("beneficiarios"):
        return "<div style='padding:20px;color:#888;font-size:13px;'>Sin datos de despacho para este proyecto.</div>"

    titulo   = datos["titulo"]
    spi_str  = datos["spi_str"]
    meses    = datos["meses"] or ["Mes 1", "Mes 2", "Mes 3"]
    bens     = datos["beneficiarios"]
    spi_col  = _spi_color(spi_str)

    # Cabecera del tab
    header = (
        f'<div style="background:#1e293b;border-radius:8px 8px 0 0;'
        f'padding:12px 18px;margin-bottom:0;">'
        f'<div style="font-size:12px;font-weight:700;color:#f1f5f9;">{titulo}</div>'
        f'<div style="font-size:11px;color:#94a3b8;margin-top:3px;">'
        f'{datos["meta"]}&nbsp;&nbsp;'
        f'<span style="background:{spi_col};color:#fff;border-radius:4px;'
        f'padding:1px 8px;font-weight:700;font-size:11px;">{spi_str}</span>'
        f'</div></div>'
    )

    # Encabezados de tabla
    ths_mes = "".join(
        f'<th style="padding:7px 8px;background:#334155;color:#cbd5e1;'
        f'font-size:10px;text-align:center;white-space:nowrap;">{m}</th>'
        for m in meses
    )
    thead = (
        '<thead><tr style="background:#1e293b;">'
        '<th style="padding:7px 12px;background:#1e293b;color:#94a3b8;'
        'font-size:10px;text-align:left;min-width:160px;">Beneficiario</th>'
        '<th style="padding:7px 8px;background:#334155;color:#cbd5e1;'
        'font-size:10px;text-align:center;white-space:nowrap;">Capataz</th>'
        '<th style="padding:7px 8px;background:#334155;color:#cbd5e1;'
        'font-size:10px;text-align:center;">Av. Viv%</th>'
        '<th style="padding:7px 8px;background:#334155;color:#cbd5e1;'
        'font-size:10px;text-align:center;">Av. Total%</th>'
        '<th style="padding:7px 8px;background:#334155;color:#cbd5e1;'
        'font-size:10px;text-align:center;">Desp. Real.</th>'
        + ths_mes +
        '<th style="padding:7px 8px;background:#1e293b;color:#94a3b8;'
        'font-size:10px;text-align:center;">P50 Días</th>'
        '</tr></thead>'
    )

    # Filas de datos agrupadas por capataz
    filas_html = ""
    capataz_actual = None
    for i, b in enumerate(bens):
        if b["capataz"] != capataz_actual:
            capataz_actual = b["capataz"]
            grupo_label = f"{b['grupo']} — {b['capataz']}" if b["grupo"] else b["capataz"]
            filas_html += (
                f'<tr><td colspan="{5 + len(meses) + 1}" style="'
                f'padding:6px 12px;background:#f1f5f9;font-size:11px;'
                f'font-weight:600;color:#334155;border-top:2px solid #e2e8f0;">'
                f'{grupo_label}</td></tr>'
            )

        bg = "#ffffff" if i % 2 == 0 else "#f8fafc"
        av_col  = _avance_color(b["av_viv"])
        at_col  = _avance_color(b["av_total"])

        celdas_mes = "".join(
            f'<td style="padding:5px 8px;border-bottom:1px solid #e2e8f0;">'
            f'{_formatear_etapas(b[k])}</td>'
            for k in ["mes1", "mes2", "mes3"]
        )

        p50 = b["p50"]
        p50_display = "—" if p50 in ("—", "", None) else p50
        try:
            p50_col = "#dc2626" if int(float(p50)) > 30 else "#16a34a"
        except Exception:
            p50_col = "#6b7280"

        filas_html += (
            f'<tr style="background:{bg};">'
            f'<td style="padding:5px 12px;font-size:11px;color:#111827;'
            f'border-bottom:1px solid #e2e8f0;">{b["nombre"]}</td>'
            f'<td style="padding:5px 8px;font-size:10px;color:#6b7280;'
            f'text-align:center;border-bottom:1px solid #e2e8f0;white-space:nowrap;">{b["capataz"]}</td>'
            f'<td style="padding:5px 8px;font-size:12px;font-weight:700;'
            f'color:{av_col};text-align:center;border-bottom:1px solid #e2e8f0;">{b["av_viv"]}%</td>'
            f'<td style="padding:5px 8px;font-size:12px;font-weight:700;'
            f'color:{at_col};text-align:center;border-bottom:1px solid #e2e8f0;">{b["av_total"]}%</td>'
            f'<td style="padding:5px 8px;font-size:12px;font-weight:600;'
            f'text-align:center;border-bottom:1px solid #e2e8f0;">{b["desp_real"]}</td>'
            + celdas_mes +
            f'<td style="padding:5px 8px;font-size:12px;font-weight:700;'
            f'color:{p50_col};text-align:center;border-bottom:1px solid #e2e8f0;">{p50_display}</td>'
            f'</tr>'
        )

    # Leyenda SOL / MC
    leyenda = (
        '<div style="display:flex;gap:12px;flex-wrap:wrap;padding:10px 12px;'
        'background:#f8fafc;border-top:1px solid #e2e8f0;font-size:10px;color:#6b7280;">'
        '<span><span style="background:#ccfbf1;color:#0f766e;border-radius:3px;'
        'padding:1px 6px;font-weight:600;">[SOL]</span> Solicitud confirmada en soldepacho</span>'
        '<span><span style="background:#fef3c7;color:#92400e;border-radius:3px;'
        'padding:1px 6px;font-weight:600;">[MC]</span> Proyección Monte Carlo</span>'
        '<span>P50 = mediana de días restantes · Rojo &gt; 30 días</span>'
        '</div>'
    )

    tabla = (
        '<div style="overflow-x:auto;border:1px solid #e2e8f0;border-radius:0 0 8px 8px;">'
        f'<table style="border-collapse:collapse;width:100%;min-width:700px;">'
        f'{thead}<tbody>{filas_html}</tbody></table>'
        f'{leyenda}</div>'
    )

    resumen_etapas = _resumen_por_etapa(bens, meses)

    return header + tabla + resumen_etapas


# ── Inyección en el HTML ──────────────────────────────────────────────────────

SENTINEL_START = "<!-- despachos-inject-start -->"
SENTINEL_END   = "<!-- despachos-inject-end -->"

DASHBOARD_PATH = Path(__file__).parents[1] / "dashboard" / "index_live_v3.html"


def _construir_bloque_script(despachos_por_idx: dict) -> str:
    """Genera el bloque JS completo (con sentinelas) para inyectar en el HTML."""
    despachos_json = json.dumps(despachos_por_idx, ensure_ascii=False)
    return (
        f"\n{SENTINEL_START}\n"
        "<script>\n"
        f"(function(){{var _d={despachos_json};"
        "Object.keys(_d).forEach(function(k){{if(secciones[k])secciones[k].despachos=_d[k];}});"
        "}})()\n"
        f"</script>\n"
        f"{SENTINEL_END}\n"
    )


def _aplicar_inyeccion(html: str, despachos_por_idx: dict) -> str:
    """
    Aplica (o actualiza) el tab Despachos en un HTML.
    Idempotente: usa centinelas para reemplazar inyecciones previas.
    """
    import re

    # 1) Botón de tab (idempotente)
    TAB_BTN = "<button class=\"tab-btn\" onclick=\"mostrar('viviendas')\">Viviendas</button>"
    TAB_BTN_NEW = (
        "<button class=\"tab-btn\" onclick=\"mostrar('viviendas')\">Viviendas</button>\n"
        "<button class=\"tab-btn\" onclick=\"mostrar('despachos')\">Despachos</button>"
    )
    if TAB_BTN in html and TAB_BTN_NEW not in html:
        html = html.replace(TAB_BTN, TAB_BTN_NEW, 1)

    # 2) Section div (idempotente)
    SEC_ANCHOR = '<div class="section" id="sec-viviendas">'
    if SEC_ANCHOR in html and "sec-despachos" not in html:
        html = html.replace(
            SEC_ANCHOR,
            '<div class="section" id="sec-despachos"></div>\n' + SEC_ANCHOR,
            1,
        )

    # 3) Bloque de datos JS (idempotente via sentinelas)
    bloque = _construir_bloque_script(despachos_por_idx)
    if SENTINEL_START in html:
        # Reemplazar inyección previa
        html = re.sub(
            re.escape(SENTINEL_START) + r".*?" + re.escape(SENTINEL_END),
            bloque.strip(),
            html,
            flags=re.DOTALL,
        )
    else:
        html = html.replace("</body></html>", bloque + "</body></html>", 1)

    return html


def _cargar_datos_proyectos() -> dict | None:
    """Descarga el Excel de Drive y retorna {pid: datos} o None si falla."""
    print("  [Despachos] Descargando Excel de Drive...")
    try:
        data = _descargar_excel()
    except Exception as e:
        print(f"  [Despachos] ERROR descargando Excel: {e}")
        return None

    print("  [Despachos] Leyendo pestañas de proyecto...")
    try:
        return _leer_todos_proyectos(data)
    except Exception as e:
        print(f"  [Despachos] ERROR leyendo Excel: {e}")
        return None


def _generar_despachos_por_idx(proyectos: dict) -> dict:
    resultado = {}
    for idx, pid in IDX_TO_PID.items():
        datos = proyectos.get(pid)
        resultado[idx] = _generar_html_proyecto(datos)
        n = len(datos["beneficiarios"]) if datos else 0
        print(f"  [Despachos] {pid} (idx {idx}): {n} beneficiarios")
    return resultado


# ── API pública ───────────────────────────────────────────────────────────────

def inyectar_resumen_despachos(pdf_dir: Path, fecha: str) -> bool:
    """Inyecta el tab Despachos en el Informe_Adquisiciones_{fecha}.html generado."""
    pdf_dir    = Path(pdf_dir)
    candidatos = sorted(pdf_dir.glob(f"Informe_Adquisiciones_{fecha}*.html"))
    if not candidatos:
        print(f"  [Despachos] No se encontró Informe_Adquisiciones_{fecha}*.html — omitiendo")
        return False

    html_path = candidatos[0]
    proyectos = _cargar_datos_proyectos()
    if proyectos is None:
        return False

    despachos_por_idx = _generar_despachos_por_idx(proyectos)
    html = _aplicar_inyeccion(
        html_path.read_text(encoding="utf-8", errors="replace"),
        despachos_por_idx,
    )
    html_path.write_text(html, encoding="utf-8")
    print(f"  [Despachos] Tab inyectado → {html_path.name} ({html_path.stat().st_size:,} bytes)")
    return True


def inyectar_en_dashboard() -> bool:
    """
    Descarga el Excel de Drive e inyecta (o actualiza) el tab Despachos en
    SCRaices-LLM/dashboard/index_live_v3.html dentro del repositorio local.
    Retorna True si el archivo cambió.
    """
    dash = DASHBOARD_PATH
    if not dash.exists():
        print(f"  [Dashboard] No encontrado: {dash} — omitiendo")
        return False

    proyectos = _cargar_datos_proyectos()
    if proyectos is None:
        return False

    despachos_por_idx = _generar_despachos_por_idx(proyectos)
    html_original = dash.read_text(encoding="utf-8", errors="replace")
    html_nuevo    = _aplicar_inyeccion(html_original, despachos_por_idx)

    if html_nuevo == html_original:
        print("  [Dashboard] Sin cambios en el dashboard.")
        return False

    dash.write_text(html_nuevo, encoding="utf-8")
    print(f"  [Dashboard] index_live_v3.html actualizado ({dash.stat().st_size:,} bytes)")
    return True
